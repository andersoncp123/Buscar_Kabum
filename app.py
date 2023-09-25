from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl

produto_busca = "4060"

# entre no site https://www.kabum.com.br/
driver = webdriver.Chrome()
driver.get("https://www.kabum.com.br/")
sleep(5)

# digita o produto_busca
campo_busca = driver.find_element(By.XPATH,"//input[@id='input-busca']")
campo_busca.send_keys(produto_busca)
sleep(1)
campo_busca.send_keys(Keys.ENTER)

# caminho dos dados
produtos_caminho_nome = driver.find_elements(By.XPATH,"//main[@class='sc-72859b2d-12 gcZYxg']//a//span[@class='sc-d79c9c3f-0 nlmfp sc-6ac6cf23-16 kWpNqb nameCard']")
produtos_caminho_preco = driver.find_elements(By.XPATH,"//main[@class='sc-72859b2d-12 gcZYxg']//a//span[@class='sc-3b515ca1-2 chPrxA priceCard']")
produtos_caminho_link = driver.find_elements(By.XPATH,"//main[@class='sc-72859b2d-12 gcZYxg']//a")

produtos_nomes = []
produtos_precos = []
produtos_links = []

for produto in produtos_caminho_nome:
    # extrai o nome e salva na lista
    produtos_nomes.append(produto.text)

for produto in produtos_caminho_preco:
    # extrai o valor e salva na lista
    produtos_precos.append(produto.text)

for produto in produtos_caminho_link:
    # extrai o link e salva na lista
    produtos_links.append(produto.get_attribute("href"))

# Para caso não tenha o arquivo "dados.xlsx" 
try:
    workbook = openpyxl.load_workbook("dados.xlsx")
except Exception as error:
    wb = openpyxl.Workbook()
    wb.save(filename='dados.xlsx')
    workbook = openpyxl.load_workbook("dados.xlsx")

if produto_busca in workbook.sheetnames:
    workbook.remove(workbook[produto_busca])

# adiciona os produtos no excel
try:
    pagina_produto = workbook[produto_busca]
    pagina_produto["A1"].value = "Nome do Produto"
    pagina_produto["B1"].value = "Preço do Produto"
    pagina_produto["C1"].value = "Link do Produto"

    # guarda no excel
    for index, rows in enumerate(pagina_produto.iter_rows(min_row = 2, max_row = len(produtos_nomes), min_col = 1, max_col = 1)):
        for cell in rows:
            cell.value = produtos_nomes[index]

    for index, rows in enumerate(pagina_produto.iter_rows(min_row = 2, max_row = len(produtos_precos), min_col = 2, max_col = 2)):
        for cell in rows:
            cell.value = produtos_precos[index]

    for index, rows in enumerate(pagina_produto.iter_rows(min_row = 2, max_row = len(produtos_links), min_col = 3, max_col = 3)):
        for cell in rows:
            cell.hyperlink = produtos_links[index]

    workbook.save("dados.xlsx")
    driver.close()

except Exception as error:
    # cria e adiciona no excel
    workbook.create_sheet(produto_busca)

    pagina_produto = workbook[produto_busca]
    pagina_produto["A1"].value = "Nome do Produto"
    pagina_produto["B1"].value = "Preço do Produto"
    pagina_produto["C1"].value = "Link do Produto"

    for index, rows in enumerate(pagina_produto.iter_rows(min_row = 2, max_row = len(produtos_nomes), min_col = 1, max_col = 1)):
        for cell in rows:
            cell.value = produtos_nomes[index]

    for index, rows in enumerate(pagina_produto.iter_rows(min_row = 2, max_row = len(produtos_precos), min_col = 2, max_col = 2)):
        for cell in rows:
            cell.value = produtos_precos[index]

    for index, rows in enumerate(pagina_produto.iter_rows(min_row = 2, max_row = len(produtos_links), min_col = 3, max_col = 3)):
        for cell in rows:
            cell.hyperlink = produtos_links[index]
    
    
workbook.save("dados.xlsx")
driver.close()
